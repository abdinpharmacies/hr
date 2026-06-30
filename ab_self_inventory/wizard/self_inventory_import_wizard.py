import base64
import io

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SelfInventoryImportWizard(models.TransientModel):
    _name = 'ab_self_inventory_import_wizard'
    _description = 'Self Inventory Actual Count Import'

    process_id = fields.Many2one('ab_self_inventory_process', required=True)
    file = fields.Binary(required=True)
    filename = fields.Char()

    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("openpyxl is required to import Excel files."))
        if self.process_id.state not in ('draft', 'in_progress'):
            raise ValidationError(_("Only active self inventory processes can import actual counts."))

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(base64.b64decode(self.file)), data_only=True)
        except Exception as exc:
            raise UserError(_("Could not read Excel file: %s") % exc)

        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValidationError(_("The Excel file is empty."))
        headers = {str(value or '').strip().lower(): index for index, value in enumerate(header_row)}
        code_index = self._first_header(headers, 'product code', 'e-plus item code', 'item code')
        actual_index = self._first_header(headers, 'actual qty', 'actual quantity')
        system_qty_index = self._first_header(headers, 'system qty', 'e-stock qty')
        if code_index is None or actual_index is None:
            raise ValidationError(_("Excel must contain Product Code and Actual Qty columns."))

        lines_by_code = {}
        for line in self.process_id.line_ids:
            for code in (line.product_code, line.eplus_item_code, str(line.eplus_item_id or '')):
                if code:
                    lines_by_code[str(code).strip()] = line

        updated = 0
        missing_codes = []
        seen_codes = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code = str(row[code_index] or '').strip()
            if not code:
                continue
            if code in seen_codes:
                raise ValidationError(_("Duplicate product code in Excel: %s") % code)
            seen_codes.add(code)
            line = lines_by_code.get(code)
            if not line:
                missing_codes.append(code)
                continue
            if system_qty_index is not None and row[system_qty_index] not in (None, ''):
                system_qty = self._to_float(row[system_qty_index], code, 'system')
                if abs(system_qty - line.system_qty) > 0.0001:
                    raise ValidationError(_("System quantity cannot be changed for product code %s.") % code)
            line.write({'actual_qty': self._to_float(row[actual_index], code, 'actual')})
            updated += 1

        if not updated:
            raise ValidationError(_("No matching inventory lines were updated."))
        message = _("Updated %s inventory lines.") % updated
        if missing_codes:
            message += _(" Missing product codes: %s") % ', '.join(missing_codes[:10])
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {'title': _('Import Complete'), 'message': message, 'type': 'success', 'sticky': False},
        }

    def _first_header(self, headers, *names):
        for name in names:
            if name in headers:
                return headers[name]
        return None

    def _to_float(self, value, code, column_type):
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            if column_type == 'system':
                raise ValidationError(_("System quantity must be numeric for product code %s.") % code)
            raise ValidationError(_("Actual quantity must be numeric for product code %s.") % code)


class SelfInventoryBatchAddLineWizard(models.TransientModel):
    _name = 'ab_self_inventory_batch_add_line_wizard'
    _description = 'Self Inventory Manual Add Line'

    request_id = fields.Many2one('ab_self_inventory_request', readonly=True)
    batch_id = fields.Many2one('ab_self_inventory_request_batch', readonly=True)
    branch_ids = fields.Many2many(
        'ab_store',
        string='Branches',
        domain="[('store_type', '=', 'branch')]",
    )
    product_ids = fields.Many2many('ab_product', string='Products', required=True)
    available_product_ids = fields.Many2many(
        'ab_product',
        compute='_compute_available_product_ids',
    )
    note = fields.Char()

    @api.depends('request_id.line_ids.product_id', 'batch_id.line_ids.product_id', 'branch_ids')
    def _compute_available_product_ids(self):
        Product = self.env['ab_product'].sudo().with_context(active_test=False)
        all_products = Product.search([])
        for wizard in self:
            used_products = self.env['ab_product']
            if wizard.request_id:
                used_products = wizard.request_id.line_ids.mapped('product_id')
            elif wizard.batch_id:
                branches = wizard.branch_ids or wizard.batch_id.branch_ids
                lines = wizard.batch_id.line_ids
                if branches:
                    lines = lines.filtered(lambda line: line.branch_id in branches)
                used_products = lines.mapped('product_id')
            wizard.available_product_ids = all_products - used_products

    def action_add_lines(self):
        self.ensure_one()
        if self.request_id:
            self._add_request_lines()
        elif self.batch_id:
            self._add_batch_lines()
        else:
            raise ValidationError(_("Open Add Line from a self inventory request or batch."))
        return {'type': 'ir.actions.act_window_close'}

    def _add_request_lines(self):
        request = self.request_id
        if request.state != 'draft':
            raise ValidationError(_("You cannot add lines after the request is submitted."))
        if not self.product_ids:
            raise ValidationError(_("Select at least one product."))
        line_values = []
        existing_products = set(request.line_ids.mapped('product_id').ids)
        duplicate_products = self.product_ids.filtered(lambda product: product.id in existing_products)
        if duplicate_products:
            raise ValidationError(
                _("These products already exist on this request: %s")
                % ', '.join(duplicate_products.mapped('display_name')[:10])
            )
        for product in self.product_ids:
            line_values.append({
                'request_id': request.id,
                'product_id': product.id,
                'eplus_item_id': int(product.eplus_serial or 0),
                'eplus_item_code': product.code or '',
                'system_qty': 0.0,
                'matched_by': 'code' if product.code else 'none',
                'selected': True,
                'sell_price': product.default_price or 0.0,
                'note': self.note,
            })
        if not line_values:
            raise ValidationError(_("Selected products already exist on this request."))
        self.env['ab_self_inventory_request_line'].create(line_values)

    def _add_batch_lines(self):
        batch = self.batch_id
        if batch.state != 'draft':
            raise ValidationError(_("You cannot add lines after the batch is submitted."))
        if not self.product_ids:
            raise ValidationError(_("Select at least one product."))
        branches = self.branch_ids or batch.branch_ids
        if not branches:
            raise ValidationError(_("Select at least one branch."))
        missing_branches = branches - batch.branch_ids
        if missing_branches:
            batch.write({'branch_ids': [(4, branch.id) for branch in missing_branches]})
        line_values = []
        existing_keys = {
            (line.branch_id.id, line.product_id.id)
            for line in batch.line_ids
            if line.branch_id and line.product_id
        }
        duplicate_names = []
        for branch in branches:
            for product in self.product_ids:
                if (branch.id, product.id) in existing_keys:
                    duplicate_names.append("%s / %s" % (branch.display_name, product.display_name))
        if duplicate_names:
            raise ValidationError(
                _("These products already exist in the result table: %s")
                % ', '.join(duplicate_names[:10])
            )
        for branch in branches:
            for product in self.product_ids:
                key = (branch.id, product.id)
                existing_keys.add(key)
                line_values.append({
                    'batch_id': batch.id,
                    'branch_id': branch.id,
                    'product_id': product.id,
                    'eplus_item_id': int(product.eplus_serial or 0),
                    'eplus_item_code': product.code or '',
                    'system_qty': 0.0,
                    'matched_by': 'code' if product.code else 'none',
                    'selected': True,
                    'sell_price': product.default_price or 0.0,
                    'note': self.note,
                })
        if not line_values:
            raise ValidationError(_("Selected products already exist for the selected branches."))
        self.env['ab_self_inventory_request_batch_line'].create(line_values)
