import re
import openpyxl

TAX_MAP = {
    'دفعات مقدمة': 'advance_payment',
    'ضريبي': 'withholding_tax',
    'غير ضريبي': 'non_taxable',
}

REGION_MAP = {
    'القاهرة': 'north',
    'الصعيد': 'south',
}

SECTION_MAP = {
    'ادويـــــــــــــــــــــــــــة': 'medicine',
    'تجميــــــــــــــــــــــل': 'cosmetics',
    'مستحضرات طبية': 'medical_preps',
    'مــــــســـتـــــــــلــــــزمـــــــات': 'supplies',
    'مـــــســــتـــــورد_أدويـــــــــــــة': 'import_medicine',
    'مـــــســــتـــــورد_تــــجــــمـــيـــل': 'import_cosmetics',
}

_STRIP_PREFIX = re.compile(r'^[#@*!]+')
_STRIP_SUFFIX = re.compile(r'[_\-]\d+.*$')
_STRIP_CODE = re.compile(r'\s+\d+.*$')
_STRIP_AT = re.compile(r'@\w+')
_STRIP_STARS = re.compile(r'\*\*\d+')
_STRIP_BRACKET = re.compile(r'[()]')
_STRIP_UDER = re.compile(r'[_\-]+')
_MULTI_SPACE = re.compile(r'\s+')
_LATIN = re.compile(r'[a-zA-Z]')


def _clean(s):
    s = _STRIP_PREFIX.sub('', s)
    s = _STRIP_SUFFIX.sub('', s)
    s = _STRIP_CODE.sub('', s)
    s = _STRIP_AT.sub('', s)
    s = _STRIP_STARS.sub('', s)
    s = _STRIP_BRACKET.sub(' ', s)
    s = s.replace('تحت التصريف', '')
    s = s.replace('تحت_التصريف', '')
    s = s.replace('_ت التصريف', '')
    s = s.replace('-تحت التصريف', '')
    s = s.replace('تحت التصريف_', '')
    s = s.replace('(تحت التصريف)', '')
    s = s.replace('\u0647', '\u0629')
    s = s.replace('\u0649', '\u064a')
    s = s.replace('\u0625', '\u0627')
    s = s.replace('\u0623', '\u0627')
    s = _STRIP_UDER.sub(' ', s)
    s = _MULTI_SPACE.sub(' ', s)
    return s.strip()


def _significant_words(text):
    words = text.split()
    result = []
    for w in words:
        w = w.strip()
        if len(w) <= 1:
            continue
        if _LATIN.search(w):
            continue
        result.append(w)
    return result


def _find_supplier(Supplier, name):
    name = name.strip()
    n = _clean(name)
    n_words = _significant_words(n)
    n_set = set(n_words)
    n_flat = ''.join(n_words)
    if not n_words:
        return Supplier.browse()
    best_score = -1
    best = Supplier.browse()
    for s in Supplier.search([]):
        sn = _clean(s.name)
        if not sn:
            continue
        sn_words = _significant_words(sn)
        sn_set = set(sn_words)
        sn_flat = ''.join(sn_words)
        common = n_set & sn_set
        has_strong_rel = (
            sn == n or sn_flat == n_flat or
            sn.startswith(n) or
            (n.startswith(sn) and len(sn_words) >= 2) or
            sn_flat.startswith(n_flat) or
            (n_flat.startswith(sn_flat) and len(sn_words) >= 2)
        )
        if not common and not has_strong_rel:
            continue
        coverage = len(common) / max(len(n_set), 1)
        if not has_strong_rel and len(common) < 2 and coverage < 0.6:
            continue
        extras = len(sn_set - n_set)
        missing = len(n_set - sn_set)
        score = coverage * 10 - extras * 0.5 - missing * 0.5
        if has_strong_rel:
            score = max(score, 12)
        if score > best_score:
            best_score = score
            best = s
    if best and best_score >= 2:
        return best
    return Supplier.browse()


_NOISE_WORDS = {'مستورد', 'ليميتد', 'فرع'}


def _strip_noise(name):
    words = name.split()
    words = [w for w in words if w not in _NOISE_WORDS]
    return ' '.join(words)


def replace_supplier_mappings(env, filepath):
    Supplier = env['ab_costcenter']
    Mapping = env['ab.supplier.mapping']
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # --- 1. Clear old data ---
    existing = Mapping.search([])
    print(f"Deleting {len(existing)} existing mapping records...")
    existing.unlink()
    # Also clear fields on any supplier that has mapping data
    tagged = Supplier.search([
        '|', ('supplier_type', '!=', False),
        '|', ('region', '!=', False),
        ('section', '!=', False),
    ])
    print(f"Clearing fields on {len(tagged)} suppliers...")
    tagged.write({'supplier_type': False, 'region': False, 'section': False})

    # --- 2. Import ---
    stats = {'matched': 0, 'skipped_no_match': 0, 'skipped_invalid': 0, 'multi_row_accounted': 0}

    for row in range(2, ws.max_row + 1):
        sname = ws.cell(row, 2).value
        tax_raw = ws.cell(row, 3).value
        section_raw = ws.cell(row, 4).value
        region_raw = ws.cell(row, 5).value

        if not sname or not tax_raw or not region_raw:
            stats['skipped_invalid'] += 1
            continue

        supplier_type = TAX_MAP.get(tax_raw.strip())
        region = REGION_MAP.get(region_raw.strip())
        section = SECTION_MAP.get(section_raw.strip()) if section_raw else False
        if not supplier_type or not region:
            stats['skipped_invalid'] += 1
            continue

        supplier = _find_supplier(Supplier, sname)
        if not supplier:
            stripped = _strip_noise(sname.strip())
            if stripped and stripped != sname.strip():
                supplier = _find_supplier(Supplier, stripped)
        if not supplier:
            words = sname.strip().split()
            for i in [0, -1]:
                if len(words) > 1:
                    trimmed = ' '.join(w for j, w in enumerate(words) if j != i)
                    supplier = _find_supplier(Supplier, trimmed)
                    if supplier:
                        break

        if not supplier:
            stats['skipped_no_match'] += 1
            continue

        vals = {'supplier_type': supplier_type, 'region': region}
        if section:
            vals['section'] = section
        supplier.write(vals)

        if not Mapping.search([('supplier_id', '=', supplier.id)], limit=1):
            Mapping.create({'supplier_id': supplier.id})

        stats['matched'] += 1

    return stats


def run(env):
    res = replace_supplier_mappings(
        env,
        '/home/abdin_02/Mohamed_tips/work/Supplier Claim/بيانات موردين.xlsx'
    )
    print("RESULT:", res)
    Mapping = env['ab.supplier.mapping']
    print(f"Total mapping records: {Mapping.search_count([])}")
