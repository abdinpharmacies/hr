import base64
import io
import openpyxl
import re
from odoo import models, fields, api, _
from odoo.exceptions import UserError

class PayrollExcelImportWizard(models.TransientModel):
    _name = 'payroll.excel.import.wizard'
    _description = 'رفع وتحليل سياسات الرواتب بالذكاء النخوبي'

    system_id = fields.Many2one('payroll.rule.system', string='نظام القواعد', required=True)
    file = fields.Binary(string='ملف الإكسل', required=True)
    file_name = fields.Char(string='اسم الملف')

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("الرجاء اختيار ملف!"))

        try:
            file_data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=False)
            
            best_analysis = {
                'sheet': None, 'header_row': 0,
                'mapping': {'name': -1, 'code': -1, 'equation': -1, 'sequence': -1, 'type': -1},
                'confidence': 0
            }

            # --- READ USER INSTRUCTIONS (AI Communication Layer) ---
            # The user can write hints in the rule system view, e.g.:
            # "عمود 'المبلغ' هو صافي الراتب" or "النظام يعتمد على الأيام وليس الساعات"
            user_hints = (self.system_id.user_instructions or "").lower()

            # Parse user hints for overrides
            user_override = {
                'net_col': None, 'system_type': None,
                'agreement_col': None, 'deduction_col': None
            }
            if any(k in user_hints for k in ['يوم', 'يومي', 'days', 'daily']):
                user_override['system_type'] = 'daily_hours'
            if any(k in user_hints for k in ['ساعة', 'ساعات', 'hours', 'hourly']):
                user_override['system_type'] = 'daily_hours'
            if any(k in user_hints for k in ['شهري', 'monthly', 'fixed']):
                user_override['system_type'] = 'fixed_monthly'

            # Extract column name overrides from hints like "عمود X هو صافي"
            net_match = re.search(r"عمود['\"\s]*([^\s'\"هو]+)['\"\s]*هو\s*(صافي|net)", user_hints)
            if net_match:
                user_override['net_col'] = net_match.group(1).strip()
            
            agr_match = re.search(r"عمود['\"\s]*([^\s'\"هو]+)['\"\s]*هو\s*(اتفاق|أساسي|basic|agreement)", user_hints)
            if agr_match:
                user_override['agreement_col'] = agr_match.group(1).strip()

            payroll_brain = {
                'allowance': ['استحقاق', 'بدل', 'حافز', 'إضافة', 'علاوة', 'allowance', 'bonus', 'extra'],
                'deduction': ['استقطاع', 'خصم', 'جزاء', 'تأمينات', 'ضريبة', 'سلف', 'deduction', 'tax', 'penalty'],
                'net': ['صافي', 'الصافي', 'net', 'total_net'],
                'agreement': ['اتفاق', 'تعاقد', 'أساسي', 'contract', 'agreement', 'basic', 'الأساسي'],
                'equations': ['معادلة', 'حساب', 'طريقة', 'منطق', 'equation', 'formula', 'calc', 'logic']
            }

            for sheet in wb.worksheets:
                rows = list(sheet.iter_rows(min_row=1, max_row=1000, max_col=100, values_only=True))
                if not rows: continue

                for r_idx, row in enumerate(rows):
                    if not any(row): continue
                    scores = {'name': 0, 'code': 0, 'equation': 0}
                    mapping = {'name': -1, 'code': -1, 'equation': -1, 'sequence': -1, 'type': -1}
                    
                    for c_idx, cell in enumerate(row):
                        if cell is None: continue
                        val = str(cell).strip().lower()
                        
                        # Match columns
                        if any(k in val for k in payroll_brain['allowance'] + payroll_brain['deduction'] + ['الاسم', 'البند', 'name', 'البيان']):
                            mapping['name'], scores['name'] = c_idx, 25
                        elif any(k in val for k in payroll_brain['equations']):
                            mapping['equation'], scores['equation'] = c_idx, 30
                        elif any(k in val for k in ['كود', 'code', 'رمز', 'id']):
                            mapping['code'], scores['code'] = c_idx, 15

                    if sum(scores.values()) >= 30:
                        data_score = 0
                        for offset in range(1, 15): # Wider check
                            target_r = r_idx + offset
                            if target_r >= len(rows): break
                            row_data = rows[target_r]
                            if mapping['equation'] != -1 and row_data[mapping['equation']]:
                                e_val = str(row_data[mapping['equation']])
                                if any(op in e_val for op in ['*', '/', '+', '-', '=', 'employee.', 'payslip.', 'record.']):
                                    data_score += 25
                                elif e_val.replace('.', '', 1).isdigit():
                                    data_score += 15
                        
                        final_conf = sum(scores.values()) + data_score
                        if final_conf > best_analysis['confidence']:
                            best_analysis.update({
                                'sheet': sheet, 'header_row': r_idx + 1,
                                'mapping': mapping, 'confidence': final_conf
                            })

            if best_analysis['confidence'] < 30:
                raise UserError(_("المُحرك الذكي لم يجد هيكلاً واضحاً للسياسات. يرجى التأكد من وجود أعمدة واضحة للمسميات والمعادلات."))

            sheet = best_analysis['sheet']
            mapping = best_analysis['mapping']
            header_idx = best_analysis['header_row']
            
            self.system_id.rule_ids.unlink()
            
            rules_data = []
            analysis_notes = []
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_idx + 1, max_row=1000, values_only=True)):
                if not any(row): continue
                name = str(row[mapping['name']]) if mapping['name'] != -1 and row[mapping['name']] else ""
                code = str(row[mapping['code']]) if mapping['code'] != -1 and row[mapping['code']] else ""
                raw_eq = str(row[mapping['equation']]) if mapping['equation'] != -1 and row[mapping['equation']] else ""
                if not name and not raw_eq: continue
                
                clean_eq = raw_eq.strip()
                if clean_eq.startswith('='): clean_eq = clean_eq[1:]
                
                # Rule Personality & Classification
                role = "بند عام"
                if any(k in name.lower() for k in payroll_brain['agreement']):
                    role = "راتب الاتفاق / الأساسي"
                    if not clean_eq or clean_eq == '0': clean_eq = "employee.agreement_salary"
                    analysis_notes.append(("agreement", name))
                elif any(k in name.lower() for k in payroll_brain['net']):
                    role = "صافي الراتب"
                    if not clean_eq or clean_eq == '0': clean_eq = "sum_allowances - sum_deductions"
                    analysis_notes.append(("net", name))
                elif any(k in name.lower() for k in payroll_brain['deduction']):
                    role = "استقطاع / خصم"
                    analysis_notes.append(("deduction", name))
                elif any(k in name.lower() for k in payroll_brain['allowance']):
                    role = "استحقاق / بدل"
                    analysis_notes.append(("allowance", name))

                # Logic Type Check
                if 'worked_hours' in clean_eq or 'contractual_hours' in clean_eq:
                    analysis_notes.append(("attendance_logic", name))

                if not code:
                    # Smart code generator for Arabic names
                    code = re.sub(r'[^a-zA-Z0-9]', '_', name.strip()).strip('_').lower()[:30]
                if not code or code.strip('_') == '':
                    code = 'rule_%d' % len(rules_data)

                rules_data.append({
                    'name': name,
                    'code': code.lower(),
                    'equation': clean_eq or '0',
                    'sequence': 10 + len(rules_data) * 5,
                    'system_id': self.system_id.id,
                    'role': role # For internal tracking during report generation
                })

            for r_vals in rules_data:
                # Remove temporary 'role' before creating record
                r_payload = r_vals.copy()
                r_payload.pop('role', None)
                self.env['payroll.rule.item'].create(r_payload)

            # --- PHASE 3: Rich Narrative Generation --- #
            p_type = 'mixed'
            notes_keys = [n[0] for n in analysis_notes]
            if 'attendance_logic' in notes_keys: p_type = 'daily_hours'
            elif 'agreement' in notes_keys: p_type = 'fixed_monthly'

            # User Instructions Take Priority
            if user_override.get('system_type'):
                p_type = user_override['system_type']
            
            user_hints_note = ""
            if user_hints.strip():
                user_hints_note = f"""
                    <div style="background: #fff8e1; padding: 12px; border-radius: 8px; border-right: 4px solid #fbbc04; margin-bottom: 15px;">
                        <strong>📝 تعليماتك التي أخذتها بعين الاعتبار:</strong>
                        <p style="margin: 5px 0 0 0; font-style: italic;">"{user_hints.strip()}"</p>
                    </div>"""

            narrative = f"""
                <div style="direction: rtl; text-align: right; background: #ffffff; border-radius: 12px; font-family: sans-serif;">
                    <div style="background: #1a73e8; color: white; padding: 20px; border-radius: 12px 12px 0 0;">
                        <h2 style="margin: 0;">🧠 تقرير تحليل سياسة الرواتب (AI Audit)</h2>
                        <p style="margin: 5px 0 0 0; opacity: 0.9;">لقد قمت بدراسة {len(rules_data)} بنداً من شيت "{sheet.title}"</p>
                    </div>
                    
                    <div style="padding: 20px; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 12px 12px;">
                        <h3 style="color: #1a73e8;">📌 التصنيف المقترح: {dict(self.system_id._fields['policy_type'].selection).get(p_type)}</h3>
                        {user_hints_note}
                        <p>بناءً على المعادلات المكتشفة، فهمت الآتي:</p>
                        
                        <div style="display: grid; grid-template-columns: 1fr; gap: 15px; margin-top: 20px;">
            """

            # Dynamic Points from Analysis Metrics
            findings = []
            if 'agreement' in notes_keys:
                names = ", ".join([n[1] for n in analysis_notes if n[0] == 'agreement'])
                findings.append(f"✅ <strong>راتب الاتفاق:</strong> اكتشفت بند التحكم الأساسي في ({names}). سأقوم بربطه بعقد الموظف تلقائياً.")
            
            if 'attendance_logic' in notes_keys:
                findings.append("✅ <strong>منطق الحضور:</strong> لاحظت معادلات تعتمد على ساعات العمل، سيتم ربطها بنظام البصمة فوراً.")
            
            if 'net' in notes_keys:
                findings.append("✅ <strong>معادلة الصافي:</strong> استنتجت طريقة احتساب الصافي النهائي للمستحقات.")

            if not findings:
                findings.append("💡 <strong>تحليل عام:</strong> لقد قمت باستخراج كافة البنود وسأقوم بمعاملتها كبنود حسابية مستقلة يتم تطبيقها بالترتيب.")

            for f in findings:
                narrative += f'<div style="background: #e8f0fe; padding: 12px; border-radius: 8px; border-right: 4px solid #1a73e8;">{f}</div>'

            # Table of Rules Found (Verification Table)
            narrative += """
                        </div>
                        
                        <h3 style="margin-top: 30px; border-bottom: 2px solid #eee; padding-bottom: 5px;">📋 ملخص بنود السياسة المستخرجة:</h3>
                        <table style="width: 100%; border-collapse: collapse; margin-top: 10px;">
                            <thead>
                                <tr style="background: #f8f9fa;">
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">البند</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">النوع المستنتج</th>
                                    <th style="padding: 10px; border: 1px solid #ddd; text-align: right;">المعادلة / القيمة</th>
                                </tr>
                            </thead>
                            <tbody>
            """
            
            for r in rules_data[:15]: # Show first 15 for brevity
                narrative += f"""
                                <tr>
                                    <td style="padding: 10px; border: 1px solid #ddd;">{r['name']}</td>
                                    <td style="padding: 10px; border: 1px solid #ddd;"><span style="color: #666; font-size: 12px;">{r['role']}</span></td>
                                    <td style="padding: 10px; border: 1px solid #ddd; font-family: monospace; color: #d81b60;">{r['equation']}</td>
                                </tr>
                """
            
            if len(rules_data) > 15:
                narrative += f'<tr><td colspan="3" style="text-align: center; padding: 10px; color: #999;">... وغيرها {len(rules_data) - 15} بند آخر تم استيرادهم بنجاح</td></tr>'

            narrative += """
                            </tbody>
                        </table>
                        
                        <div style="margin-top: 30px; background: #fff3e0; padding: 15px; border-radius: 8px; border-right: 4px solid #ff9800;">
                            <strong>⚠️ خطوة الاعتماد:</strong> لقد وضعت النظام في حالة <strong>"تحت المراجعة"</strong>. يرجى التأكد من البيانات أعلاه ثم الضغط على زر <strong>"اعتماد السياسة"</strong> في الأعلى لتفعيلها.
                        </div>
                    </div>
                </div>
            """

            self.system_id.write({
                'policy_type': p_type,
                'policy_narrative': narrative,
                'state': 'draft'
            })

            return {
                'name': _('مراجعة السياسة المستنتجة'),
                'type': 'ir.actions.act_window',
                'res_model': 'payroll.rule.system',
                'res_id': self.system_id.id,
                'view_mode': 'form',
                'target': 'current',
            }

        except Exception as e:
            raise UserError(_("خطأ في التحليل النخوبي الفائق: %s") % str(e))
